import openpyxl
wb = openpyxl.load_workbook('C:/Users/aizen/Desktop/Pangea POS Data.xlsx')
ws = wb['Sheet1']
print('Rows:', ws.max_row, '| Cols:', ws.max_column)
print('Headers:', [c.value for c in ws[1]])
print()
for row in ws.iter_rows(min_row=2, max_row=8):
    print([c.value for c in row])
